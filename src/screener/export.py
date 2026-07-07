import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def calculate_composite_score(df):
    """
    Calculate composite quality score (0–100).
    """

    df = df.copy()

    score = (
        df["roe"] * 0.35
        + df["fcf"] * 0.30
        + df["revenue_cagr_5yr"] * 0.20
        + (100 - df["de"]) * 0.15
    )

    maximum = score.max()

    if maximum == 0:
        df["composite_quality_score"] = 0
    else:
        df["composite_quality_score"] = (
            score / maximum * 100
        ).round(2)

    return df


def export_to_excel(df, filename="output/screener_output.xlsx"):
    """
    Export dataframe to Excel.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Screener Results"

    ws.append(list(df.columns))

    green = PatternFill(
        start_color="90EE90",
        end_color="90EE90",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= 50:
                    cell.fill = green
                else:
                    cell.fill = red

    wb.save(filename)

    return filename


def export_screener(df, filename="output/screener_output.xlsx"):
    """
    Calculate score and export.
    """

    df = calculate_composite_score(df)

    return export_to_excel(df, filename)